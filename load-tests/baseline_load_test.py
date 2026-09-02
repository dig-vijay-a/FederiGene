"""
FederiGene - Baseline / Load Testing Script
=============================================
Simulates 100 concurrent virtual users hitting the API continuously for 1 minute.

Metrics collected:
  - Requests per second (RPS)
  - Response Time: Average, Min, Max, P50, P95, P99
  - Error rate
  - Status code distribution
  - Per-endpoint breakdown

Usage:
  pip install aiohttp openpyxl
  python baseline_load_test.py                         # Default: localhost:8000
  python baseline_load_test.py --base-url http://staging.federigene.com
  python baseline_load_test.py --users 200 --duration 120
"""

import asyncio
import aiohttp
import time
import statistics
import json
import sys
import argparse
from collections import defaultdict
from datetime import datetime


# ============================================================================
# CONFIGURATION
# ============================================================================
DEFAULT_BASE_URL = "http://localhost:8000"
DEFAULT_VIRTUAL_USERS = 100
DEFAULT_DURATION_SECONDS = 60

# Endpoints to test (method, path, body_or_none, name)
ENDPOINTS = [
    # Public / Health
    ("GET",  "/health",                           None, "Health Check"),
    ("GET",  "/",                                 None, "Root / SPA"),

    # Auth - Public
    ("GET",  "/api/auth/security-questions",      None, "Get Security Questions"),
    ("POST", "/api/auth/check-availability",      {"email": "loadtest@example.com", "username": "loadtest_user"}, "Check Availability"),
    ("POST", "/api/auth/login",                   {"identifier": "loadtest@fake.com", "password": "WrongPassword123!"}, "Login Attempt (Expected 401)"),
    ("GET",  "/api/auth/pre-reg-totp",            None, "Pre-Registration TOTP"),

    # License - Public
    ("GET",  "/api/license/tiers",                None, "License Tiers"),

    # Compliance - Public
    ("GET",  "/api/compliance/frameworks",        None, "Compliance Frameworks"),

    # Marketplace - Public
    ("GET",  "/api/marketplace/catalog",          None, "Marketplace Catalog"),
    ("GET",  "/api/marketplace/models",           None, "Marketplace Models"),
]


# ============================================================================
# METRICS COLLECTOR
# ============================================================================
class MetricsCollector:
    def __init__(self):
        self.lock = asyncio.Lock()
        self.response_times = []                      # All response times (ms)
        self.per_endpoint = defaultdict(list)          # Per-endpoint response times
        self.status_codes = defaultdict(int)           # Status code distribution
        self.per_endpoint_status = defaultdict(lambda: defaultdict(int))
        self.errors = 0
        self.total_requests = 0
        self.total_bytes = 0
        self.start_time = None
        self.end_time = None
        self.timeline = []                             # (timestamp, latency_ms) for RPS calc

    async def record(self, endpoint_name, status_code, latency_ms, response_bytes, is_error=False):
        async with self.lock:
            self.total_requests += 1
            self.response_times.append(latency_ms)
            self.per_endpoint[endpoint_name].append(latency_ms)
            self.status_codes[status_code] += 1
            self.per_endpoint_status[endpoint_name][status_code] += 1
            self.total_bytes += response_bytes
            self.timeline.append((time.time(), latency_ms))
            if is_error:
                self.errors += 1

    def get_summary(self):
        if not self.response_times:
            return {"error": "No requests were completed"}

        duration = self.end_time - self.start_time
        sorted_times = sorted(self.response_times)
        n = len(sorted_times)

        # Calculate RPS over 1-second windows
        rps_windows = defaultdict(int)
        for ts, _ in self.timeline:
            second = int(ts - self.start_time)
            rps_windows[second] += 1

        rps_values = list(rps_windows.values()) if rps_windows else [0]

        return {
            "test_config": {
                "virtual_users": DEFAULT_VIRTUAL_USERS,
                "duration_seconds": duration,
                "endpoints_tested": len(ENDPOINTS),
                "timestamp": datetime.now().isoformat(),
            },
            "throughput": {
                "total_requests": self.total_requests,
                "total_data_transferred_kb": round(self.total_bytes / 1024, 2),
                "rps_average": round(self.total_requests / duration, 2),
                "rps_max": max(rps_values),
                "rps_min": min(rps_values) if len(rps_values) > 2 else min(rps_values),
            },
            "response_time_ms": {
                "average": round(statistics.mean(sorted_times), 2),
                "min": round(sorted_times[0], 2),
                "max": round(sorted_times[-1], 2),
                "median_p50": round(sorted_times[n // 2], 2),
                "p90": round(sorted_times[int(n * 0.9)], 2),
                "p95": round(sorted_times[int(n * 0.95)], 2),
                "p99": round(sorted_times[int(n * 0.99)], 2),
                "stdev": round(statistics.stdev(sorted_times), 2) if n > 1 else 0,
            },
            "reliability": {
                "success_count": self.total_requests - self.errors,
                "error_count": self.errors,
                "error_rate_percent": round((self.errors / self.total_requests) * 100, 2),
            },
            "status_codes": dict(self.status_codes),
        }

    def get_per_endpoint_summary(self):
        results = []
        for name, times in self.per_endpoint.items():
            if not times:
                continue
            sorted_t = sorted(times)
            n = len(sorted_t)
            status_dist = dict(self.per_endpoint_status[name])

            # Count errors (5xx or connection errors)
            endpoint_errors = sum(v for k, v in status_dist.items() if k >= 500 or k == 0)

            results.append({
                "endpoint": name,
                "total_requests": n,
                "avg_ms": round(statistics.mean(sorted_t), 2),
                "min_ms": round(sorted_t[0], 2),
                "max_ms": round(sorted_t[-1], 2),
                "p50_ms": round(sorted_t[n // 2], 2),
                "p95_ms": round(sorted_t[int(n * 0.95)], 2),
                "p99_ms": round(sorted_t[int(n * 0.99)], 2),
                "error_rate": round((endpoint_errors / n) * 100, 2),
                "status_codes": status_dist,
            })

        return sorted(results, key=lambda x: x["avg_ms"], reverse=True)


# ============================================================================
# VIRTUAL USER (Worker)
# ============================================================================
async def virtual_user(user_id, base_url, metrics, stop_event):
    """
    Simulates a single virtual user continuously making requests
    until the stop_event is set.
    """
    timeout = aiohttp.ClientTimeout(total=10)
    connector = aiohttp.TCPConnector(limit=0, force_close=False)

    async with aiohttp.ClientSession(timeout=timeout, connector=connector) as session:
        while not stop_event.is_set():
            for method, path, body, name in ENDPOINTS:
                if stop_event.is_set():
                    break

                url = f"{base_url}{path}"
                start = time.perf_counter()
                status_code = 0
                resp_bytes = 0
                is_error = False

                try:
                    if method == "GET":
                        params = {}
                        if path == "/api/auth/pre-reg-totp":
                            params = {"email": f"user{user_id}@loadtest.com"}
                        async with session.get(url, params=params) as resp:
                            status_code = resp.status
                            data = await resp.read()
                            resp_bytes = len(data)
                    elif method == "POST":
                        headers = {"Content-Type": "application/json"}
                        async with session.post(url, json=body, headers=headers) as resp:
                            status_code = resp.status
                            data = await resp.read()
                            resp_bytes = len(data)

                    # 4xx are expected (e.g., 401 on login) — not counted as errors
                    if status_code >= 500:
                        is_error = True

                except asyncio.TimeoutError:
                    status_code = 0
                    is_error = True
                except aiohttp.ClientError:
                    status_code = 0
                    is_error = True
                except Exception:
                    status_code = 0
                    is_error = True

                elapsed_ms = (time.perf_counter() - start) * 1000
                await metrics.record(name, status_code, elapsed_ms, resp_bytes, is_error)


# ============================================================================
# LIVE PROGRESS DISPLAY
# ============================================================================
async def progress_display(metrics, duration, stop_event):
    """Shows live stats every 5 seconds during the test."""
    start = time.time()
    while not stop_event.is_set():
        await asyncio.sleep(5)
        elapsed = time.time() - start
        remaining = max(0, duration - elapsed)

        async with metrics.lock:
            total = metrics.total_requests
            errs = metrics.errors
            if metrics.response_times:
                avg_rt = round(statistics.mean(metrics.response_times[-500:]), 1)
            else:
                avg_rt = 0

        rps = round(total / elapsed, 1) if elapsed > 0 else 0
        bar_len = 30
        progress = min(elapsed / duration, 1.0)
        filled = int(bar_len * progress)
        bar = "#" * filled + "-" * (bar_len - filled)

        print(
            f"\r  [{bar}] {elapsed:.0f}s/{duration}s | "
            f"Reqs: {total:,} | RPS: {rps} | "
            f"Avg: {avg_rt}ms | Errors: {errs}",
            end="", flush=True
        )

    print()  # newline after progress


# ============================================================================
# MAIN TEST RUNNER
# ============================================================================
async def run_load_test(base_url, num_users, duration):
    metrics = MetricsCollector()
    stop_event = asyncio.Event()

    print("=" * 70)
    print("  FederiGene — Baseline / Load Test")
    print("=" * 70)
    print(f"  Target:          {base_url}")
    print(f"  Virtual Users:   {num_users}")
    print(f"  Duration:        {duration} seconds")
    print(f"  Endpoints:       {len(ENDPOINTS)}")
    print(f"  Started:         {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    print()

    # Pre-flight check
    print("  [1/4] Pre-flight connectivity check...", end=" ", flush=True)
    try:
        timeout = aiohttp.ClientTimeout(total=5)
        async with aiohttp.ClientSession(timeout=timeout) as s:
            async with s.get(f"{base_url}/health") as r:
                if r.status == 200:
                    print("OK")
                else:
                    print(f"WARNING (status {r.status})")
    except Exception as e:
        print(f"FAILED ({e})")
        print()
        print("  ERROR: Cannot connect to the backend. Make sure the server is running:")
        print(f"         cd Backend && uvicorn main:app --host 0.0.0.0 --port 8000")
        print()
        return None

    # Ramp up
    print(f"  [2/4] Ramping up {num_users} virtual users...", flush=True)
    metrics.start_time = time.time()

    # Create virtual user tasks
    user_tasks = []
    for i in range(num_users):
        task = asyncio.create_task(virtual_user(i, base_url, metrics, stop_event))
        user_tasks.append(task)

    # Create progress display task
    progress_task = asyncio.create_task(progress_display(metrics, duration, stop_event))

    print(f"  [3/4] Load test running for {duration} seconds...")
    print()

    # Wait for test duration
    await asyncio.sleep(duration)

    # Signal stop
    stop_event.set()
    metrics.end_time = time.time()

    # Wait for all users to finish current requests
    print("  [4/4] Cooling down, waiting for in-flight requests...")
    await asyncio.sleep(2)

    # Cancel remaining tasks
    for task in user_tasks:
        task.cancel()
    progress_task.cancel()

    try:
        await asyncio.gather(*user_tasks, return_exceptions=True)
    except Exception:
        pass

    return metrics


# ============================================================================
# CONSOLE REPORT
# ============================================================================
def print_report(metrics):
    summary = metrics.get_summary()
    per_ep = metrics.get_per_endpoint_summary()

    print()
    print("=" * 70)
    print("  LOAD TEST RESULTS")
    print("=" * 70)

    # Throughput
    tp = summary["throughput"]
    print()
    print("  --- Throughput ---")
    print(f"  Total Requests:       {tp['total_requests']:,}")
    print(f"  Data Transferred:     {tp['total_data_transferred_kb']:,.1f} KB")
    print(f"  Requests/sec (avg):   {tp['rps_average']}")
    print(f"  Requests/sec (peak):  {tp['rps_max']}")

    # Response Time
    rt = summary["response_time_ms"]
    print()
    print("  --- Response Time ---")
    print(f"  Average:    {rt['average']} ms")
    print(f"  Min:        {rt['min']} ms")
    print(f"  Max:        {rt['max']} ms")
    print(f"  Median:     {rt['median_p50']} ms")
    print(f"  P90:        {rt['p90']} ms")
    print(f"  P95:        {rt['p95']} ms")
    print(f"  P99:        {rt['p99']} ms")
    print(f"  Std Dev:    {rt['stdev']} ms")

    # Reliability
    rel = summary["reliability"]
    print()
    print("  --- Reliability ---")
    print(f"  Successful:   {rel['success_count']:,}")
    print(f"  Errors:       {rel['error_count']:,}")
    print(f"  Error Rate:   {rel['error_rate_percent']}%")

    # Status Codes
    print()
    print("  --- Status Code Distribution ---")
    for code, count in sorted(summary["status_codes"].items()):
        label = {200: "OK", 401: "Unauthorized", 404: "Not Found", 422: "Validation Error", 0: "Connection Error/Timeout"}.get(code, "Other")
        print(f"    {code:>4}: {count:>8,} ({label})")

    # Per-Endpoint Breakdown
    print()
    print("  --- Per-Endpoint Breakdown ---")
    print(f"  {'Endpoint':<30} {'Reqs':>7} {'Avg(ms)':>8} {'Min':>7} {'Max':>8} {'P95':>8} {'Err%':>6}")
    print("  " + "-" * 78)
    for ep in per_ep:
        print(
            f"  {ep['endpoint']:<30} "
            f"{ep['total_requests']:>7,} "
            f"{ep['avg_ms']:>8.1f} "
            f"{ep['min_ms']:>7.1f} "
            f"{ep['max_ms']:>8.1f} "
            f"{ep['p95_ms']:>8.1f} "
            f"{ep['error_rate']:>5.1f}%"
        )

    print()
    print("=" * 70)


# ============================================================================
# EXCEL REPORT GENERATOR
# ============================================================================
def generate_excel_report(metrics, output_path="load-test-report.xlsx"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARN] openpyxl not installed. Skipping Excel report.")
        print("         Install with: pip install openpyxl")
        return

    summary = metrics.get_summary()
    per_ep = metrics.get_per_endpoint_summary()

    wb = Workbook()

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # ===================== SHEET 1: Executive Summary =====================
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = "2F5496"

    ws1.merge_cells("A1:B1")
    ws1["A1"] = "FederiGene Load Test - Executive Summary"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=16, color="2F5496")

    config = summary["test_config"]
    tp = summary["throughput"]
    rt = summary["response_time_ms"]
    rel = summary["reliability"]

    rows = [
        ("", ""),
        ("TEST CONFIGURATION", ""),
        ("Virtual Users", config["virtual_users"]),
        ("Duration (seconds)", round(config["duration_seconds"], 1)),
        ("Endpoints Tested", config["endpoints_tested"]),
        ("Test Timestamp", config["timestamp"]),
        ("", ""),
        ("THROUGHPUT", ""),
        ("Total Requests", tp["total_requests"]),
        ("Data Transferred (KB)", tp["total_data_transferred_kb"]),
        ("Requests/sec (Average)", tp["rps_average"]),
        ("Requests/sec (Peak)", tp["rps_max"]),
        ("Requests/sec (Min)", tp["rps_min"]),
        ("", ""),
        ("RESPONSE TIME (ms)", ""),
        ("Average", rt["average"]),
        ("Minimum", rt["min"]),
        ("Maximum", rt["max"]),
        ("Median (P50)", rt["median_p50"]),
        ("P90", rt["p90"]),
        ("P95", rt["p95"]),
        ("P99", rt["p99"]),
        ("Standard Deviation", rt["stdev"]),
        ("", ""),
        ("RELIABILITY", ""),
        ("Successful Requests", rel["success_count"]),
        ("Failed Requests", rel["error_count"]),
        ("Error Rate (%)", rel["error_rate_percent"]),
        ("", ""),
        ("STATUS CODE DISTRIBUTION", ""),
    ]

    for code, count in sorted(summary["status_codes"].items()):
        rows.append((f"HTTP {code}", count))

    for i, (label, value) in enumerate(rows, start=2):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=value)
        if label in ("TEST CONFIGURATION", "THROUGHPUT", "RESPONSE TIME (ms)", "RELIABILITY", "STATUS CODE DISTRIBUTION"):
            ws1.cell(row=i, column=1).font = section_font

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 25

    # Color-code the error rate
    err_row = None
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=2):
        if row[0].value == "Error Rate (%)":
            err_row = row[1]
            break
    if err_row:
        if rel["error_rate_percent"] == 0:
            err_row.fill = good_fill
        elif rel["error_rate_percent"] < 5:
            err_row.fill = warn_fill
        else:
            err_row.fill = bad_fill

    # ===================== SHEET 2: Per-Endpoint Breakdown =====================
    ws2 = wb.create_sheet("Endpoint Breakdown")
    ws2.sheet_properties.tabColor = "548235"

    headers = ["Endpoint", "Total Requests", "Avg (ms)", "Min (ms)", "Max (ms)",
               "P50 (ms)", "P95 (ms)", "P99 (ms)", "Error Rate (%)"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header_row(ws2, 1, len(headers))

    for i, ep in enumerate(per_ep, start=2):
        ws2.cell(row=i, column=1, value=ep["endpoint"])
        ws2.cell(row=i, column=2, value=ep["total_requests"])
        ws2.cell(row=i, column=3, value=ep["avg_ms"])
        ws2.cell(row=i, column=4, value=ep["min_ms"])
        ws2.cell(row=i, column=5, value=ep["max_ms"])
        ws2.cell(row=i, column=6, value=ep["p50_ms"])
        ws2.cell(row=i, column=7, value=ep["p95_ms"])
        ws2.cell(row=i, column=8, value=ep["p99_ms"])
        ws2.cell(row=i, column=9, value=ep["error_rate"])

        # Color-code response time
        avg_cell = ws2.cell(row=i, column=3)
        if ep["avg_ms"] < 100:
            avg_cell.fill = good_fill
        elif ep["avg_ms"] < 500:
            avg_cell.fill = warn_fill
        else:
            avg_cell.fill = bad_fill

        for col in range(1, len(headers) + 1):
            ws2.cell(row=i, column=col).border = thin_border

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    ws2.column_dimensions["A"].width = 32

    # ===================== SHEET 3: RPS Timeline =====================
    ws3 = wb.create_sheet("RPS Timeline")
    ws3.sheet_properties.tabColor = "BF8F00"

    ws3.cell(row=1, column=1, value="Second")
    ws3.cell(row=1, column=2, value="Requests in Window")
    ws3.cell(row=1, column=3, value="Cumulative Requests")
    style_header_row(ws3, 1, 3)

    rps_windows = defaultdict(int)
    for ts, _ in metrics.timeline:
        second = int(ts - metrics.start_time)
        rps_windows[second] += 1

    cumulative = 0
    for sec in sorted(rps_windows.keys()):
        row = sec + 2
        cumulative += rps_windows[sec]
        ws3.cell(row=row, column=1, value=sec)
        ws3.cell(row=row, column=2, value=rps_windows[sec])
        ws3.cell(row=row, column=3, value=cumulative)

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 22

    # Add RPS chart
    if len(rps_windows) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Requests Per Second Over Time"
        chart.y_axis.title = "Requests"
        chart.x_axis.title = "Second"
        chart.style = 10
        chart.width = 25
        chart.height = 15

        data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(rps_windows) + 1)
        cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(rps_windows) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws3.add_chart(chart, "E2")

    # ===================== SHEET 4: Raw Timing Data (sampled) =====================
    ws4 = wb.create_sheet("Response Time Distribution")
    ws4.sheet_properties.tabColor = "C00000"

    ws4.cell(row=1, column=1, value="Percentile")
    ws4.cell(row=1, column=2, value="Response Time (ms)")
    style_header_row(ws4, 1, 2)

    sorted_times = sorted(metrics.response_times)
    n = len(sorted_times)
    percentiles = [1, 5, 10, 25, 50, 75, 90, 95, 99, 99.5, 99.9, 100]

    for i, p in enumerate(percentiles, start=2):
        idx = min(int(n * p / 100), n - 1)
        ws4.cell(row=i, column=1, value=f"P{p}")
        ws4.cell(row=i, column=2, value=round(sorted_times[idx], 2))
        ws4.cell(row=i, column=1).border = thin_border
        ws4.cell(row=i, column=2).border = thin_border

    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["B"].width = 22

    # ===================== SAVE =====================
    wb.save(output_path)
    print(f"  [OK] Excel report saved: {output_path}")


# ============================================================================
# ENTRY POINT
# ============================================================================
def main():
    parser = argparse.ArgumentParser(description="FederiGene Baseline/Load Test")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL, help="Backend base URL")
    parser.add_argument("--users", type=int, default=DEFAULT_VIRTUAL_USERS, help="Number of virtual users")
    parser.add_argument("--duration", type=int, default=DEFAULT_DURATION_SECONDS, help="Test duration in seconds")
    parser.add_argument("--output", default="load-test-report.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    # DEFAULT_VIRTUAL_USERS = args.users

    metrics = asyncio.run(run_load_test(args.base_url, args.users, args.duration))

    if metrics and metrics.total_requests > 0:
        print_report(metrics)
        generate_excel_report(metrics, args.output)

        # Also save JSON summary
        json_path = args.output.replace(".xlsx", ".json")
        with open(json_path, "w") as f:
            json.dump(metrics.get_summary(), f, indent=2)
        print(f"  [OK] JSON summary saved: {json_path}")
        print()
    else:
        print("  No data collected. Ensure the backend is running.")


if __name__ == "__main__":
    main()
